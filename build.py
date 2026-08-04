from crewai import crew
import sys
from rich import default_styles
from rich import default_styles
import os
import csv
import subprocess
import yaml
from pathlib import Path
import re

import openpyxl


# 멀티라인 문자열을 리터럴 블록 스칼라(|) 스타일로 출력하는 커스텀 Dumper
class LiteralDumper(yaml.Dumper):
    pass

def _str_representer(dumper, data):
    """개행(\n)이 포함된 문자열은 | 블록 스타일로 출력합니다."""
    if '\n' in data:
        # 각 줄의 후행 공백을 제거 (PyYAML은 줄 끝 공백이 있으면 | 스타일 거부)
        lines = data.split('\n')
        data = '\n'.join(line.rstrip() for line in lines)
        data = data.rstrip()
        return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='|')
    return dumper.represent_scalar('tag:yaml.org,2002:str', data)

LiteralDumper.add_representer(str, _str_representer)


def xlsx_to_csv(xlsx_path, csv_path=None, sheet_index=0):
    """
    xlsx 파일의 지정된 시트를 csv 파일로 변환합니다.
    
    Args:
        xlsx_path: 입력 xlsx 파일 경로
        csv_path: 출력 csv 파일 경로 (None이면 확장자만 .csv로 변경)
        sheet_index: 변환할 시트 인덱스 (기본값: 0 = 첫 번째 시트)
    
    Returns:
        생성된 csv 파일 경로
    """
    if csv_path is None:
        csv_path = os.path.splitext(xlsx_path)[0] + '.csv'

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if sheet_index >= len(sheet_names):
        wb.close()
        raise IndexError(
            f"시트 인덱스 {sheet_index}가 범위를 벗어났습니다. "
            f"사용 가능한 시트: {sheet_names}"
        )

    ws = wb[sheet_names[sheet_index]]
    print(f"[Convert] '{xlsx_path}' → '{csv_path}' (시트: '{sheet_names[sheet_index]}')")

    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            # None 값을 빈 문자열로 치환
            writer.writerow([cell if cell is not None else '' for cell in row])

    wb.close()
    print(f"[Success] CSV 변환 완료. ({ws.max_row}행 × {ws.max_column}열)")
    return csv_path


def build_yaml_configs(csv_file_path):
    """CSV 스펙을 읽어 YAML 설정용 딕셔너리로 변환합니다."""
    crews_config = {}

    with open(csv_file_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        crews_config = {
            'agents': {}, 'tasks': {}, 
            'crewai_tools': set(), # crewai_tools 제공 툴
            'custom_tools': set(),   # 자체 제작 툴
            'agent_tools': {}
        }
        for row in reader:

            agent_key = row['task_agent']
            task_key = row.get('task_name', f"task_{agent_key}") # task_name이 없으면 자동 생성
            custom_tool = row.get('custom_tool', '')
            crewai_tool = row.get('crewai_tool', '')

            if agent_key not in crews_config['agent_tools']:
                crews_config['agent_tools'][agent_key] = set()

            for tool in custom_tool.split(','):
                if tool != '' :
                    crews_config['agent_tools'][agent_key].add(tool)
                    crews_config['custom_tools'].add(tool)

            for tool in crewai_tool.split(','):
                if tool != '' :
                    crews_config['agent_tools'][agent_key].add(tool)
                    crews_config['crewai_tools'].add(tool)

            # Agents 설정 구성
            if agent_key not in crews_config['agents']:
                crews_config['agents'][agent_key] = {
                    'role': row['persona_role'],
                    'goal': row['persona_goal'],
                    'backstory': row['persona_backstory']
                }

            # Tasks 설정 구성
            crews_config['tasks'][task_key] = {
                'description': row['task_description'],
                'expected_output': row['task_expected_output'],
                'agent': agent_key,
                #'context': row['task_context']
            }
            if row['task_context'] != '':
                # 쉼표로 구분된 복수 context 지원, YAML 리스트로 저장
                context_list = [c.strip() for c in row['task_context'].split(',') if c.strip()]
                crews_config['tasks'][task_key]['context'] = context_list

    return crews_config


def run_scaffolding(crew_name):
    """CrewAI CLI를 사용하여 기본 프로젝트 스캐폴딩을 생성합니다."""
    print(f"\n[Scaffolding] '{crew_name}' 크루 생성을 시작합니다.")
    try:
        # 기존 폴더가 없을 경우 자동(비대화형) 생성
        if not os.path.exists(crew_name):
            # --skip_provider 옵션으로 프롬프트 질의 자동 처리
            subprocess.run(["crewai", "create", "crew", crew_name, "--skip_provider"], check=True)
            print(f"[Success] '{crew_name}' 스캐폴딩 완료.")
            return True
        else:
            print(f"[Skip] '{crew_name}' 폴더가 이미 존재합니다.")
            return True
    except subprocess.CalledProcessError as e:
        print(f"[Error] CLI 실행 중 오류 발생: {e}")
        return False


def generate_custom_tools_file(crew_name, package_name, config):
    """분류된 Custom Tool들을 위한 Pseudo 코드를 custom_tool.py에 생성합니다."""

    custom_tools = config['custom_tools']
    
    if not custom_tools:
        return # 커스텀 툴이 없으면 스킵

    # CrewAI 스캐폴딩의 기본 도구 폴더 경로
    tool_file = Path(crew_name) / "src" / package_name / "tools" / "custom_tool.py"
    tool_file.parent.mkdir(parents=True, exist_ok=True) # tools 폴더가 없으면 생성

    # tool_file 이 기존에 존재하면 custom_tool.py.1 형태로 롤링해서 백업함
    if tool_file.exists():
        # 가장 오래된 .py.9 백업이 있다면 미리 삭제
        max_backup = tool_file.with_suffix('.py.9')
        if max_backup.exists():
            max_backup.unlink()

        # 존재하는 파일에 대해서만 .py.8 -> .py.9, .py.7 -> .py.8 로 1단계씩 시프트
        for i in range(8, 0, -1):
            src = tool_file.with_suffix(f'.py.{i}')
            dst = tool_file.with_suffix(f'.py.{i+1}')
            if src.exists():
                src.rename(dst)

        # 원본 custom_tool.py -> custom_tool.py.1 로 변경
        tool_file.rename(tool_file.with_suffix('.py.1'))

    
    # 1. 텍스트에서 {변수명} 추출 (중복 제거를 위해 set 사용)
    input_vars = set()
    for task in config['tasks'].values():
        desc_vars = re.findall(r'\{(\w+)\}', task.get('description', ''))
        exp_vars = re.findall(r'\{(\w+)\}', task.get('expected_output', ''))
        input_vars.update(desc_vars)
        input_vars.update(exp_vars)

    input_vars_code1 = ''
    for var in input_vars:
        input_vars_code1 += f'{var}: str = Field(..., description="{var}에 대한 설명을 자세하게 입력하세요.")\n    '
    #print(input_vars_code1)
    
    input_vars_code2 = ''
    for var in input_vars:
        input_vars_code2 += f'{var}:str, '
    # 마지막 , 을 제거
    input_vars_code2 = input_vars_code2.rstrip(', ')
    #print(input_vars_code2)

    # BaseTool 상속 및 Pydantic 스키마 템플릿
    content = """from crewai.tools import BaseTool
from typing import Type
from pydantic import BaseModel, Field

"""
    # 정의된 커스텀 툴 갯수만큼 클래스 생성
    for tool_name in custom_tools:
        content += f"""class {tool_name}Input(BaseModel):
    \"\"\"Input schema for {tool_name}.\"\"\"
    {input_vars_code1}

class {tool_name}(BaseTool):
    name: str = "{tool_name}"
    description: str = "{tool_name}에 대한 설명을 자세하게 입력하세요."
    args_schema: Type[BaseModel] = {tool_name}Input

    def _run(self, {input_vars_code2}) -> str:
        # [TODO: Pseudo-code] 실제 비즈니스 로직을 여기에 구현하세요.
        # 예: import pandas as pd
        # df = pd.read_csv('data.csv')
        # return df[df['target'] == query].to_string()
        
        return f"[{tool_name}] 성공적으로 실행되었습니다."

"""
    with open(tool_file, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"[Create] '{tool_file}' 에 커스텀 툴 Pseudo 코드 생성 완료.")


def update_crew_py_file(crew_name, package_name, config):
    """
    YAML 설정의 Key(agent_name, task_name)를 읽어와서
    crew.py의 데코레이터 메서드들을 동적으로 생성하고 덮어씁니다.
    """
    crew_file = Path(crew_name) / "src" / package_name / "crew.py"
    
    # 1. 클래스명 생성 (snake_case -> CamelCase 변환)
    # 예: my_auto_crew -> MyAutoCrew
    class_name = "".join([word.capitalize() for word in crew_name.replace('-', '_').split("_")])

    # ------------------------------------------------
    # [수정] Import 구문 분리 (Standard vs Custom)
    # ------------------------------------------------
    imports_code = ""

    #print('config.get("crewai_tools") : ', config.get('crewai_tools'))
    if config.get('crewai_tools'):
        tools_str = ", ".join(config['crewai_tools'])
        imports_code += f"from crewai_tools import {tools_str}\n"
        
    #print('config.get("custom_tools") : ', config.get('custom_tools'))
    if config.get('custom_tools'):
        custom_tools_str = ", ".join(config['custom_tools'])
        imports_code += f"from {package_name}.tools.custom_tool import {custom_tools_str}\n"

    # 2. Agent 메서드 코드 동적 생성
    agents_code = ""
    for agent_key in config['agents'].keys():
        assigned_tools = config['agent_tools'].get(agent_key, set())
        tools_inject_str = ""
        if assigned_tools:
            tools_list_str = ", ".join([f"{t}()" for t in assigned_tools])
            tools_inject_str = f"            tools=[{tools_list_str}],\n"

        agents_code += f"""
    @agent
    def {agent_key}(self) -> Agent:
        return Agent(
            config=self.agents_config['{agent_key}'],
{tools_inject_str}            verbose=True,
            reasoning=False,
            max_reasoning_attempts=None,
            inject_date=True,
            allow_delegation=False,
            max_iter=20,
            max_rpm=None,
            max_execution_time=None,
            llm=llm
        )
    """

    # 3. Task 메서드 코드 동적 생성
    tasks_code = ""
    for task_key in config['tasks'].keys():
        tasks_code += f"""
    @task
    def {task_key}(self) -> Task:
        return Task(
            config=self.tasks_config['{task_key}'],
            markdown=False
        )
"""

    # 4. 전체 crew.py 템플릿 결합
    # 주의: 파이썬 코드의 들여쓰기(Indentation)가 정확해야 하므로 f-string의 공백 유지
    crew_py_content = f"""from crewai import Agent, Crew, Process, Task
from crewai.project import CrewBase, agent, crew, task
{imports_code}
# devx api 호출
from .devx_llm_wrapper import llm

@CrewBase
class {class_name}():
    \"\"\"{class_name} crew generated by Meta-Agent\"\"\"

    agents_config = 'config/agents.yaml'
    tasks_config = 'config/tasks.yaml'
{agents_code}{tasks_code}
    @crew
    def crew(self) -> Crew:
        \"\"\"Creates the {class_name} crew\"\"\"
        return Crew(
            agents=self.agents, # 자동으로 @agent가 붙은 메서드들을 수집
            tasks=self.tasks,   # 자동으로 @task가 붙은 메서드들을 수집
            process=Process.sequential,
            verbose=True,
            # process=Process.hierarchical, # 관리자 에이전트 도입 등 필요시 변경
        )
"""

    # 5. 기존 파일 덮어쓰기
    if crew_file.exists():
        with open(crew_file, 'w', encoding='utf-8') as f:
            f.write(crew_py_content)
        print(f"[Update] '{crew_file}' 파일 동적 생성 및 덮어쓰기 완료.")
    else:
        print(f"[Error] '{crew_file}' 경로를 찾을 수 없습니다. 스캐폴딩이 정상적으로 생성되었는지 확인하세요.")


def update_config_files(crew_name, package_name, config):
    """생성된 스캐폴딩 내의 YAML 파일들을 스펙에 맞게 덮어씁니다."""
    # CrewAI CLI(버전 0.30+ 기준)는 src/crew_name/config/ 하위에 yaml을 생성합니다.
    # 프로젝트 이름에서 '-'를 '_'로 변환한 패키지 폴더명을 찾습니다.
    package_name = crew_name.replace('-', '_')
    config_dir = Path(crew_name) / "src" / package_name / "config"
    
    if not config_dir.exists():
        print(f"[Warning] 설정 폴더를 찾을 수 없습니다: {config_dir}")
        return

    agents_file = config_dir / "agents.yaml"
    tasks_file = config_dir / "tasks.yaml"

    # 1. agents.yaml 업데이트
    with open(agents_file, 'w', encoding='utf-8') as f:
        yaml.dump(config['agents'], f, Dumper=LiteralDumper, allow_unicode=True, sort_keys=False)
        
    # 2. tasks.yaml 업데이트
    with open(tasks_file, 'w', encoding='utf-8') as f:
        yaml.dump(config['tasks'], f, Dumper=LiteralDumper, allow_unicode=True, sort_keys=False)
        
    print(f"[Update] '{crew_name}'의 agents.yaml 및 tasks.yaml 파일 갱신 완료.")


def update_main_py_file(crew_name, package_name, config):
    """
    Task 텍스트(설명, 결과물)에서 {변수}를 추출하여
    inputs 딕셔너리가 동적으로 적용된 main.py를 생성합니다.
    """
    main_file = Path(crew_name) / "src" / package_name / "main.py"
    class_name = "".join([word.capitalize() for word in crew_name.replace('-', '_').split("_")])

    # 1. 텍스트에서 {변수명} 추출 (중복 제거를 위해 set 사용)
    input_vars = set()
    for task in config['tasks'].values():
        desc_vars = re.findall(r'\{(\w+)\}', task.get('description', ''))
        exp_vars = re.findall(r'\{(\w+)\}', task.get('expected_output', ''))
        input_vars.update(desc_vars)
        input_vars.update(exp_vars)

    # 2. inputs 딕셔너리 문자열 동적 생성
    if input_vars:
        inputs_str = "    inputs = {\n"
        for var in input_vars:
            inputs_str += f"        '{var}': '여기에 {var} 값을 입력하세요', # TODO: 런타임에 실제 값으로 치환\n"
        inputs_str += "    }"
    else:
        inputs_str = "    inputs = {} # CSV 텍스트에 설정된 동적 변수({var})가 없습니다."

    # 3. main.py 전체 템플릿 코드
    main_py_content = f"""#!/usr/bin/env python
import sys
from {package_name}.crew import {class_name}

# This main file is automatically generated by the Meta-Agent.

def run():
    \"\"\"Run the crew.\"\"\"
{inputs_str}
    try:
        {class_name}().crew().kickoff(inputs=inputs)
    except Exception as e:
        raise Exception(f"An error occurred while running the crew: {{e}}")

"""

    # 4. 파일 덮어쓰기
    if main_file.exists() or main_file.parent.exists():
        with open(main_file, 'w', encoding='utf-8') as f:
            f.write(main_py_content)
        print(f"[Update] '{main_file}' 파일(동적 inputs 딕셔너리 포함) 덮어쓰기 완료.")
    else:
        print(f"[Error] '{main_file}' 경로를 찾을 수 없습니다.")


def add_necessary_files(crew_name, package_name, config):
    """생성된 스캐폴딩에 후속 파일을 추가합니다."""
    
    # 1. devx_llm_wrapper.py 복사 (to src/[CREW_NAME]/)
    target_dir = Path(crew_name) / "src" / package_name
    if not target_dir.exists():
        print(f"[Warning] 대상 폴더를 찾을 수 없습니다: {target_dir}")
    else :     
        target_file = target_dir / "devx_llm_wrapper.py"
        with open(target_file, 'w', encoding='utf-8') as f:
            f.write(open('../devx_llm_wrapper.py').read())
        print(f"[Copy] '{crew_name}'의 devx_llm_wrapper.py 파일 갱신 완료.")    
    
    # 2. custom env 내용을  .env 에 추가
    target_dir = Path(crew_name)
    if not target_dir.exists():
        print(f"[Warning] 대상 폴더를 찾을 수 없습니다: {target_dir}")
    else :     
        target_file = target_dir / ".env"
        with open(target_file, 'w', encoding='utf-8') as f:
            f.write(open('../env.local').read()) # 본인의 env.xxx 로 대체해서 사용
        print(f"[Copy] '{crew_name}'의 .env 파일 갱신 완료.")    


def main():

    # 실행 시 argument 로 파일 이름 설정
    args = sys.argv

    # for i in range(0, len(args)):
    #     print(f"argument[{i}] : {args[i]}")
    
    if len(args) < 2:
        print(f"[{args[0]}] 사용법: uv run build.py [crew_name]")
        return

    crew_name = args[1]
    xlsx_path = crew_name + ".xlsx" if crew_name is not None else ''
    
    if not os.path.exists(xlsx_path):
        print(f"[{xlsx_path}] 파일을 찾을 수 없습니다.")
        return

    # 1. XLSX를 CSV로 변환
    csv_path = xlsx_to_csv(xlsx_path)

    # 2. CSV 파싱하여 설정 데이터 추출
    crews_config = build_yaml_configs(csv_path)

    # 하위에 generated_crews 폴더를 만들고 그 하위에서 아래 작업을 한다.
    if not os.path.exists("generated_crews"):
        os.makedirs("generated_crews")
    os.chdir("generated_crews") 

    # 3. 파싱된 데이터를 바탕으로 각각의 Crew 스캐폴딩 및 파일 업데이트
    package_name = crew_name.replace('-', '_')

    # 1. CrewAI CLI를 사용하여 기본 프로젝트 스캐폴딩을 생성
    if run_scaffolding(crew_name):

        # 2. Custom Tool 코드 먼저 생성
        generate_custom_tools_file(crew_name, package_name, crews_config)

        # 3. crew.py 동적 생성 및 덮어쓰기
        update_crew_py_file(crew_name, package_name, crews_config)

        # 4. 생성된 폴더 내 YAML 파일 덮어쓰기
        update_config_files(crew_name, package_name, crews_config)

        # 5. main.py 업데이트 (동적 inputs 주입)
        update_main_py_file(crew_name, package_name, crews_config)
    
        # 6. 후속 작업
        add_necessary_files(crew_name, package_name, crews_config)

        print("\n✅ 모든 자동화 애플리케이션 생성 프로세스가 완료되었습니다.")

    else:
        print("\n❌ 자동화 애플리케이션 생성 프로세스가 중단되었습니다.")    


if __name__ == "__main__":
    main()